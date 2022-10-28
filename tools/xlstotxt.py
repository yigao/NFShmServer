#!/bin/usr/env python
# coding=utf-8
# author:wanglian



from openpyxl import load_workbook
import sys
import re
import os
from time import perf_counter_ns, perf_counter

#是否是windowns系统
is_win_system = False
if os.name == 'nt':
    is_win_system = True
    import msvcrt

#当前目录
cur_path = os.getcwd()
# excel 表格目录
xls_file_path = '../' #os.path.dirname(cur_path)
# 项目目录
proj_path = xls_file_path + '../'  #os.path.dirname(xls_file_path)
# list.txt 文件目录
list_file = xls_file_path + 'list.txt'
# 生成的txt 文件目录
txt_file_path = proj_path + 'product/server/data/config/'
# 生成的 .h 代码 文件目录
h_file_path = proj_path + 'server/code/include/Common/TableData/'
# 生成的 .cpp 代码 文件目录
cpp_file_path = proj_path + 'server/code/source/Common/TableData/'
# 分隔符
split_char = '#'


def valid_dt(dt):
    '''
    是否是有效的数据类型
    :param dt: 数据类型
    :return:
    '''
    return (dt == 'string' or dt == 'int' or dt == 'float' or dt == 'int64')

def dt_func(dt):
    func = ''
    if 'int' == dt:
        func = 'atoi'
    elif 'int64' == dt:
        func = 'atoll'
    elif 'float' == dt:
        func = 'atof'
    return func

def dt_assign_str(dt, en, idxname, comments):
    func = dt_func(dt)
    if len(func) > 0:
        return f'{en} = {func}(pRow->at({idxname}++).c_str());   //{comments}'
    return f'{en} = pRow->at({idxname}++).c_str();   //{comments}'

def dt_push_assign_str(dt, en, idxname, comments):
    func = dt_func(dt)
    if len(func) > 0:
        return f'{en}.push_back( {func}(pRow->at({idxname}++).c_str()) );   //{comments}'
    return f'{en}.push_back(pRow->at({idxname}++).c_str());   //{comments}'


def dt_format(dt):
    format = ''
    if 'int' == dt or 'int32_t' == dt:
        format = '%d'
    elif 'int64' == dt or 'int64_t' == dt:
        format = '%ld'
    elif 'float' == dt:
        format = '%f'
    elif 'string' == dt:
        format = '%s'
    return format

def dt_str(dt):
    dtstr = ''
    if 'int' == dt:
        dtstr = 'int32_t'
    elif 'int64' == dt:
        dtstr = 'int64_t'
    elif 'float' == dt:
        dtstr = 'float'
    elif 'string' == dt:
        dtstr = 'string'
    return dtstr

def dt_ary_str(dt):
    dtstr = ''
    if 'int' == dt:
        dtstr = 'VEC_INT32'
    elif 'int64' == dt:
        dtstr = 'VEC_INT64'
    elif 'float' == dt:
        dtstr = 'VEC_FLOAT'
    elif 'string' == dt:
        dtstr = 'VEC_STRING'
    return dtstr


def struct_member(dt, en, comments):
    '''
    构造结构体初始化成员的字符串
    :param dt:
    :param en:
    :param comments:
    :return:
    '''
    if 'int' == dt:
        return f'int32_t {en} = 0;     //{comments}'
    elif 'int64' == dt:
        return f'int64_t {en} = 0;     //{comments}'
    elif 'float' == dt:
        return f'float {en} = 0.0f;     //{comments}'
    elif 'string' == dt:
        return f'string {en};     //{comments}'
    else:
        return ''

def struct_ary_member(dt, en, comments):
    '''
    单个元素的数组，构造结构体成员
    :param dt:
    :param en:
    :param comments:
    :return:
    '''
    return f'{dt_ary_str(dt)}  vec_{en};  //{comments}'

def is_english_char(ch):
    '''
    是否是26个英文字符(包括大小写)
    :param ch: 字符
    :return:
    '''
    ch_code = ord(ch)
    return ( (ch_code >= 97 and ch_code <= 122)  or (ch_code >= 65 and ch_code <= 90) )



def is_digital_char(ch):
    '''
    是否是数字字符 '0' - '9'
    :param ch: 字符
    :return:
    '''
    ch_code = ord(ch)
    return ( ch_code >= 48 and ch_code <= 57 )

def is_split_char(ch):
    '''
    是否是分隔符
    :param ch:
    :return:
    '''
    return True if ch in split_char else False

def cell_str_value(sheet, row, col):
    '''
    返回表格指定行和列处的值,以字符串的形式返回
    :param sheet:
    :param row:
    :param col:
    :return:
    '''
    return '' if sheet.cell(row, col).value is None else str(sheet.cell(row, col).value)

def has_split_char(str):
    '''
    字符串是否含有分隔符
    :param str:
    :return:
    '''
    return True if split_char in str else False


def valid_english(en):
    '''
    是否是有效的英文名字
    :param en: 英文字符串
    :return:
    '''
    en_len = len(en)
    if en_len <= 0:
        return False
    for i in range(en_len):
        ch = en[i]
        if i == 0 and  not is_english_char(ch):
            return False
        if ch == '_' or is_english_char(ch) or is_digital_char(ch):
            continue
        return False
    return True


def parse_chinese(ch_str, en_str):
    '''
    解析字段中文注释, 如果注释中含有数字，表示该字段是数组
    :param ch_str:
    :return: (结果, 第几个元素, 数组元素英文key, 数组元素英文字段名, 数组元素中文key, 数组元素中文注释)
    '''
    ret_str = ch_str.strip()
    ch_len = len(ret_str)
    if ch_len <= 0:
        return (True, 0, '', '', '', '')
    #查找字符串中的数字字符串
    ret = re.search(r'\d+',ret_str)
    if ret is not None:  #如果是数组, 数组元素列注释一定要有数字, 从1开始
        digital_str = ret.group()
        ch_idx = int(digital_str)
        if ch_idx <= 0:  #错误: 数组中文注释中数字小于等于0
            return (False, 0, '', '', '', '')
        #中文注释以数字进行分割
        (pre_str, _, post_str) = ret_str.partition(digital_str)
        #英文名字以下划线进行分割
        (en_pre_str, _, en_post_str) = en_str.partition('_')
        #分割后的 前半部分 英文字符串一定不能为空
        if len(en_pre_str) <= 0:
            return (False, ch_idx, en_pre_str, en_post_str, pre_str, post_str)
        #
        if len(en_post_str) <= 0:
            en_post_str =en_pre_str
        return (True, ch_idx, en_pre_str, en_post_str, pre_str, post_str)

    return (True, 0, '', '', '', '')




class Array():
    '''
    配置中的数组信息
    '''
    def __init__(self, arykey,chkey, enlst, chlst, first):
        self._arykey = arykey   #数组的key  比如数组字段: attribute_type, attribute_valueMin, 则数组key为  'attribute'
        self._chkey = chkey     #中文key    比如: 基础属性1、基础属性1最小数值,基础属性1最大数值 的中文key为 '基础属性'
        self._en_lst = enlst    #英文字段名列表，每个数组元素包含的字段列表
        self._ch_lst = chlst    #中文字段名列表，每个数组元素包含的字段中文注释 列表
        self._count = 0         #数组中元素数量
        self._elem_cnt = 0      #上次校验到了 数组元素的 第几字段
        self._first_col =first  #数组开始的列号

    @property
    def AryKey(self):
        return self._arykey

    @property
    def Chkey(self):
        return self._chkey

    @property
    def Count(self):
        return self._count

    #每个数组元素有多少个字段
    @property
    def Fieldnum(self):
        return len(self._en_lst)

    def Enfield(self, idx):
        #指定位置的 英文字段名
        return self._en_lst[idx]

    def Chcomments(self,idx):
        #指定位置 英文字段名对应的 中文注释
        return self._ch_lst[idx]

    @property
    def Firstcol(self):
        return self._first_col

    def AddElem(self, elem_no, en_val, ch_val):
        if len(en_val) <= 0:
            return (False, f' array elem field name length is zero...en_val:{en_val}, key:{self._arykey} ')
        if elem_no <= 0:
            return (False, f' array elem_no <= 0...elem_no:{elem_no}, en_val:{en_val}, key:{self._arykey} ')

        if elem_no - 1 != self._count:
            return (False, f' array elem_no - 1 != self._count...elem_no:{elem_no}, count:{self._count}, en_val:{en_val}, key:{self._arykey} ')
        #
        if en_val != self._en_lst[self._elem_cnt]:
            return (False, f' array en field name not match...en_val:{en_val}, old_enval:{self._en_lst[self._elem_cnt]},  key:{self._arykey} ')
        if ch_val != self._ch_lst[self._elem_cnt]:
            return (False, f' array ch field name not match...en_val:{en_val}, ch_val:{ch_val}, old_chval:{self._ch_lst[self._elem_cnt]},  key:{self._arykey} ')
        self._elem_cnt = self._elem_cnt + 1
        if self._elem_cnt >= len(self._en_lst):
            self._count = self._count + 1
            self._elem_cnt = 0
        #
        #print(f'key:{self.AryKey} count:{self._count}  elem_no:{self._elem_cnt}')
        return (True, ' array add elem success...')




class Head():
    '''
    表头
    '''
    def __init__(self, en, ch, dt, colidx, arykey):
        self._en = en           #英文字段名
        self._ch = ch           #代码中字段的 中文注释
        self._dt = dt           #类型字符串
        self._colidx = colidx   #所在的列,从1开始
        self._arykey = arykey   #配置字段作为数组时表示数组唯一的key,主要用于关联外部数组信息的

    @property
    def En(self):
        return self._en

    @property
    def Ch(self):
        return self._ch

    @property
    def Arykey(self):
        return self._arykey

    @property
    def Dt(self):
        return self._dt

    @property
    def Colidx(self):
        return self._colidx




def read_sheet_head(xlsname, sheet):
    result = []
    ary_dict = {}
    max_row = sheet.max_row
    max_col = sheet.max_column
    sheetname = sheet.title
    if max_row < 4 or max_col <= 0:
        return (False, [], {} )
    #print(f'max_row:{max_row}, max_col:{max_col} ')
    for i in range(max_col):
        # 表头全部为 None
        r1 = sheet.cell(1, i + 1).value
        r2 = sheet.cell(2, i + 1).value
        r3 = sheet.cell(3, i + 1).value
        r4 = sheet.cell(4, i + 1).value
        # print(f'head{i+1}: {r1},{r2},{r3},{r4}')
        if r1 is None and r2 is None and r3 is None and r4 is None:
            if i == 0:
                print(f'head error: first col all value is None...{xlsname}_{sheetname} ')
                return (False, [], {})
            # 遇到表头全部为 None, 表示达到表最后一列
            #print(f'max_col:{i}, {len(result)}')
            break
        #
        flag = str(sheet.cell(4, i + 1).value)
        if not flag.isdecimal():
            print(f'head error: invalid flag:{flag}  col:{i+1}, {xlsname}_{sheetname} ')
            return (False, [], {})
        #
        flag = int(flag)
        if flag != 2 and flag != 3:  #标记为 2 或 3 的才需要导出
            continue
        en = str(sheet.cell(1, i+1).value)
        if len(en) <= 0: # 如果英文字段名为空,直接跳过这一列
            continue
        ch = cell_str_value(sheet, 2, i + 1)
        dt = cell_str_value(sheet, 3, i + 1)
        if not valid_english(en):
            print(f'head error: invalid english name:{en}, ch:{ch} col:{i+1}  {xlsname}_{sheetname} ')
            return (False, [], {} )
        if not valid_dt(dt):
            print(f'head error: invalid data type:{dt}, name:{en} ch:{ch}  col:{i+1}  {xlsname}_{sheetname} ')
            return (False, [], {} )
        # 解析中文注释
        (res, elem_no, en_key, en_val, ch_key, ch_val) = parse_chinese(ch, en)
        if not res:
            print(f'head error: parse_chinese error...elem_no:{elem_no}, en_key:{en_key}, en_val:{en_val}, ch_key:{ch_key},ch_val:{ch_val},  col:{i+1}  {xlsname}_{sheetname} ')
            return (False, [], {} )

        # 数组字段
        if elem_no > 0 and ary_dict.get(en_key) is None:
            # 数组对象的 key 已存在, 存在重复的 配置表英文字段名
            en_lst = [en_val]
            ch_lst = [ch_val]
            array_end = False
            for k in range(i+1, max_col):
                temp_en = cell_str_value(sheet, 1, k + 1)
                temp_ch = cell_str_value(sheet, 2, k + 1)
                temp_flag = cell_str_value(sheet, 4, k + 1)
                if not temp_flag.isdecimal():
                    print(f'head error: invalid flag:{temp_flag}  col:{k+1}, {xlsname}_{sheetname} ')
                    return (False, [], {} )
                #
                temp_flag = int(temp_flag)
                if temp_flag != 2 and temp_flag != 3:  # 标记为 2 或 3 的才需要导出
                    continue
                (temp_res, temp_elem_no, temp_en_key, temp_en_val, temp_ch_key, temp_ch_val) = parse_chinese(temp_ch, temp_en)
                if not temp_res:
                    print(f'head error: parse_chinese error..can not find next array element info...elem_no:{elem_no}, en_key:{en_key}, en_val:{en_val}, ch_key:{ch_key},ch_val:{ch_val},  col:{i+1}  {xlsname}_{sheetname} ')
                    return (False, [], {} )
                #
                if elem_no == temp_elem_no and temp_ch_key == ch_key and temp_en_key == en_key and temp_en_val not in en_lst:
                    en_lst.append(temp_en_val)
                    ch_lst.append(temp_ch_val)
                    if k == max_col - 1:
                        array_end = True
                else:
                    array_end = True
                #
                if array_end:
                    #print(f'key:{en_key}, chkey:{ch_key}, field:{en_lst}, firstcol:{len(result)} ')
                    #ary_dict[en_key] = Array(en_key,ch_key,en_lst, ch_lst, i + 1)
                    ary_dict[en_key] = Array(en_key, ch_key, en_lst, ch_lst, len(result))
                    break
        #
        ary_key = ''
        if elem_no > 0:
            ary_key = en_key
            #添加数组元素字段
            array = ary_dict.get(en_key)
            if array is None:
                print(f'head error: can not find array....ary_key:{en_key}, en_val:{en_val}, ch_key:{ch_key}, ch_val:{ch_val}, col:{i + 1}, {xlsname}->{sheetname} ')
                return (False, [], {} )
            (add_ret, add_err) = array.AddElem(elem_no, en_val, ch_val)
            if not add_ret:
                print(f'head error: array AddElem failed....err:{add_err}, col:{i + 1}, {xlsname}->{sheetname} ')
                return (False, [], {})
        #
        obj = Head(en, ch, dt, i + 1, ary_key)
        result.append(obj)
        #print(f'{ i + 1}, {en}, {ch}, {dt}, {ary_key}')

    return (True, result, ary_dict)


# 数据写入 txt
def write_txt_file(xlsname, sheet, head_lst):

    max_row = sheet.max_row
    sheetname = sheet.title
    txt_name = txt_file_path + xlsname.capitalize() + '_' + sheetname + '.txt'
    #print(f'max_row:{max_row}, max_col:{sheet.max_column}, len:{len(head_lst)} ')
    #print(f'head_len:{len(head_lst)}')
    try:
        with open(txt_name, 'w') as file:
            hdlen = len(head_lst)
            content = ''
            row_count = 0
            for r in range(4, max_row):
                col_num = 0
                row_str = ''
                for idx in range(hdlen):  # Head 对象
                    row = r + 1
                    col = head_lst[idx].Colidx #obj._colidx #obj.Colidx
                    #
                    col_num = col_num + 1
                    val = cell_str_value(sheet, row, col)
                    if col_num <= 1 and len(val) <= 0:   #第一列为key,如果为空,直接忽略
                        break
                    if has_split_char(val):
                        print(f'write_txt_file {xlsname}->{sheetname} cell:{row}, col:{col}  has {split_char} in value:{val} ')
                        row_str = ''
                        return False
                    if col_num < hdlen:
                        row_str = f'{row_str}{val}{split_char}'  #row_str + '#'
                    else:
                        row_str = f'{row_str}{val}\n' #row_str + '\n'
                    #if sheetname == 'equip':
                    #    print(f'{row}:{col}  {sheet.cell(row, col).value}->{val}')
                #
                if len(row_str) > 0:
                    content = f'{content}{row_str}'
                    row_count = row_count + 1
                    if row_count >= 50:
                        file.write(content)
                        content = ''
                        row_count = 0
            #写入文件
            if len(content) > 0:
                # 第一行写入空行
                #file.write('\n')
                # 写入剩下的内容
                file.write(content)
    except Exception as err:
        if os.path.exists(txt_name):
            os.remove(txt_name)
        print(f'write_txt_file {xlsname}->{sheetname} error : {err} ')
        return False
    print(f'\n{txt_name}')
    return True

def write_table_class_h_file(xlsname,sheetname,head_lst, ary_dict):
    '''
    表格数据结构、类定义 写入文件
    :return:
    '''
    def_file_name = f'{h_file_path}{xlsname.capitalize()}{sheetname.capitalize()}Cfg.h'
    try:
        with open(def_file_name, 'w') as f:
            f.write(f'#pragma once\n#include "base/core/singleton.h"\n#include "Common/ComTypeDefine.h"\n\n')
            lbrace = '{'
            rbrace = '}'
            return_str = 'return'
            # 数组元素 数据结构
            for obj in ary_dict.values():
                #print(f'{obj.AryKey}')
                fieldnum = obj.Fieldnum
                if fieldnum <= 1:
                    continue
                firstcol = obj.Firstcol

                dt_name = f'{xlsname.capitalize()}{sheetname.capitalize()}{obj.AryKey.capitalize()}'
                struct_name = f'{dt_name}CfgInfo'
                field_content = ''
                #print(f'111->{fieldnum}, {firstcol}')
                for i in range(fieldnum):
                    en = obj.Enfield(i)
                    comments = obj.Chcomments(i)
                    type = head_lst[firstcol + i].Dt
                    field_content = f'{field_content}     {struct_member(type,en,comments)}\n'
                #print(f'111->0')
                if len(field_content) > 0:
                    f.write(f'\nstruct {struct_name}\n'
                            f'{lbrace}\n'
                            f'{field_content}'
                            f'{rbrace};\n'
                            f'using Vec{dt_name}Cfg = vector<{struct_name}>;\n')
                #print(f'111->1')
            # 表的数据结构
            #print('222')
            dt_name = f'{xlsname.capitalize()}{sheetname.capitalize()}'
            struct_name = f'{dt_name}CfgInfo'
            field_content = ''
            ary_key_list = []  # 已经生成代码的数组key
            for elem in head_lst:
                en = elem.En
                comments = elem.Ch
                type = elem.Dt
                arykey = elem.Arykey
                aryflag = False  #是否是数组成员
                if len(arykey) > 0:
                    aryobj = ary_dict.get(arykey)
                    if aryobj is not None:
                        aryflag = True
                        if arykey in ary_key_list:
                            continue
                        #
                        ary_key_list.append(arykey)
                        if  aryobj.Fieldnum > 1:
                            ary_dt_name = f'{xlsname.capitalize()}{sheetname.capitalize()}{aryobj.AryKey.capitalize()}'
                            field_content = f'{field_content}       Vec{ary_dt_name}Cfg vec{ary_dt_name}Cfg;    //{aryobj.Chkey}\n'
                        else:
                            field_content = f'{field_content}       {struct_ary_member(type,en, comments)}\n'
                #
                if not aryflag:
                    field_content = f'{field_content}       {struct_member(type, en, comments)}\n'
            # 数据结构的定义 写入头文件
            #print('333')
            if len(field_content) > 0:
                f.write(f'\nstruct {struct_name}\n'
                        f'{lbrace}\n'
                        f'{field_content}'
                        f'{rbrace};\n'
                        f'using {dt_name}CfgMap = '
                        f'unordered_map<{dt_str(head_lst[0].Dt)}, {struct_name}>; \n\n\n')
            # 类定义 写入文件
            #print('444')
            f.write(f'class {dt_name}Cfg : public ManualSingleton <{dt_name}Cfg>\n'
                    f'{lbrace}\n'
                    f'public:\n'
                    f'      {dt_name}Cfg();\n'
                    f'      ~{dt_name}Cfg();\n'
                    f'      bool Init(const char *szCfgFile, const char *szSplitChar);\n'
                    f'      bool UnInit();\n'
                    f'      void Clear();\n'
                    f'      const {dt_name}CfgInfo *'
                    f'Get{dt_name}CfgInfo({dt_str(head_lst[0].Dt)} {head_lst[0].En});\n'
                    f'      const {dt_name}CfgMap *'
                    f'Get{dt_name}CfgMap() {lbrace} {return_str} &m_{dt_name}CfgMap; {rbrace}\n'
                    f'private:\n'
                    f'      {dt_name}CfgMap m_{dt_name}CfgMap;\n'
                    f'{rbrace};\n\n'
                    f'#define  g_Make{dt_name}CfgTable() ({dt_name}Cfg::Instance())\n'
                    f'#define  g_Get{dt_name}CfgTable()  ({dt_name}Cfg::GetInstance())\n'
                    f'#define  g_Del{dt_name}CfgTable()  ({dt_name}Cfg::Destroy())\n')

            #print('555')
    except Exception as err:
        if os.path.exists(def_file_name):
            os.remove(def_file_name)
        print(f'write_table_class_h_file {xlsname}->{sheetname} error : {err} ')
        return False

    print(f'{def_file_name}')
    return True

def write_table_class_cpp_file(xlsname, sheetname, head_lst, ary_dict):
    imp_file_name = f'{cpp_file_path}{xlsname.capitalize()}{sheetname.capitalize()}Cfg.cpp'
    dt_name = f'{xlsname.capitalize()}{sheetname.capitalize()}'
    lbrace = '{'
    rbrace = '}'
    try:
        with open(imp_file_name, 'w') as f:
            ary_key_list = []  # 已经生成代码的数组key
            field_content = ''
            objname='info'
            idxname = 'j'
            total_colnum = len(head_lst)
            for elem in head_lst:
                en = elem.En
                comments = elem.Ch
                type = elem.Dt
                arykey = elem.Arykey
                aryflag = False  #是否是数组成员
                if len(arykey) > 0:
                    aryobj = ary_dict.get(arykey)
                    if aryobj is not None:
                        aryflag = True
                        if arykey in ary_key_list:
                            continue
                        #
                        firstcol = aryobj.Firstcol
                        ary_key_list.append(arykey)
                        if  aryobj.Fieldnum > 1:
                            ary_dt_name = f'{xlsname.capitalize()}{sheetname.capitalize()}{aryobj.AryKey.capitalize()}'
                            if aryobj.Count > 0:
                                field_content = f'{field_content}' \
                                                f'for (int t=0; t<{aryobj.Count};t++)\n          ' \
                                                f'{lbrace}\n          ' \
                                                f'   {ary_dt_name}CfgInfo obj;\n          '
                                #
                                for n in range(aryobj.Fieldnum):
                                    tmp_type = head_lst[firstcol + n].Dt
                                    tmp_en = aryobj.Enfield(n)
                                    tmp_comments = aryobj.Chcomments(n)
                                    field_content = f'{field_content}   obj.{dt_assign_str(tmp_type, tmp_en, idxname, tmp_comments)}\n          '
                                #
                                field_content = f'{field_content}   {objname}.vec{ary_dt_name}Cfg.push_back(obj);\n          ' \
                                                f'{rbrace}\n          '
                        else:
                            if aryobj.Count > 0:
                                #print(f'en:{aryobj.Enfield(0)}, ch:{aryobj.Chcomments(0)}')
                                tmp_type = head_lst[firstcol].Dt
                                #tmp_en = f'vec_{aryobj.Enfield(0)}'
                                tmp_en = f'vec_{head_lst[firstcol].En}'
                                tmp_comments = aryobj.Chcomments(0)
                                field_content = f'{field_content}' \
                                                f'for (int t=0; t<{aryobj.Count};t++)\n          ' \
                                                f'{lbrace}\n          ' \
                                                f'  {objname}.{dt_push_assign_str(tmp_type, tmp_en, idxname, tmp_comments)};\n          ' \
                                                f'{rbrace}\n          '
                #
                if not aryflag:
                    field_content = f'{field_content}{objname}.{dt_assign_str(type, en, idxname, comments)}\n          '
            #
            key_dt = dt_str(head_lst[0].Dt)
            key_en = head_lst[0].En
            obj_key_menber = f'{objname}.{head_lst[0].En}'
            f.write(f'#include "Common/TableData/{dt_name}Cfg.h"\n'
                    f'#include "Common/Reader/CfgReader.h"\n'
                    f'#include "base/core/log.h"\n\n'
                    f'{dt_name}Cfg::{dt_name}Cfg()\n'
                    f'{lbrace}\n'
                    f'{rbrace}\n\n'
                    f'{dt_name}Cfg::~{dt_name}Cfg()\n'
                    f'{lbrace}\n'
                    f'{rbrace}\n\n'
                    f'void {dt_name}Cfg::Clear()\n'
                    f'{lbrace}\n'
                    f'  m_{dt_name}CfgMap.clear();\n'
                    f'{rbrace}\n\n'
                    f'bool {dt_name}Cfg::Init(const char *szCfgFile, const char *szSplitChar)\n'
                    f'{lbrace}\n'
                    f'  CfgReader reader;\n'
                    f'  if (!reader.LoadCfg(szCfgFile, szSplitChar))\n'
                    f'  {lbrace}\n'
                    f'      LogErrFmtPrint("[common] reader config error....cfg:%s ", szCfgFile);\n'
                    f'      return false;\n'
                    f'  {rbrace}\n'
                    f'  int32_t row = reader.GetRowCount();\n'
                    f'  for (int i = 0; i < row; ++i)\n'
                    f'  {lbrace}\n'
                    f'      const VecRow *pRow = reader.GetRow(i);\n'
                    f'      if (nullptr != pRow)\n'
                    f'      {lbrace}\n'
                    f'          int32_t colun = pRow->size();\n'
                    f'          int32_t j = 0;\n'
                    f'          if (colun != {total_colnum})\n'
                    f'          {lbrace}\n'
                    f'              LogErrFmtPrint("[common] config colun count error....cfg:%s,row:%d,colun:%d,count:%d ", szCfgFile,i+1,colun,{total_colnum});\n'
                    f'              continue;\n'
                    f'          {rbrace}\n\n'
                    f'          {dt_name}CfgInfo {objname};\n'
                    f'          {field_content}\n'
                    f'          if (nullptr != Get{dt_name}CfgInfo({obj_key_menber} ))\n'
                    f'          {lbrace}\n'
                    f'              LogErrFmtPrint("[common] load config error...repeated key...cfg:%s, key:{dt_format(key_dt)}  ", szCfgFile,{obj_key_menber});\n'
                    f'              return false;\n'
                    f'          {rbrace}\n'
                    f'          m_{dt_name}CfgMap[{obj_key_menber}] = {objname};\n'
                    f'      {rbrace}\n'
                    f'  {rbrace}\n'
                    f'  return true;\n'
                    f'{rbrace}\n\n'
                    f'bool {dt_name}Cfg::UnInit()\n'
                    f'{lbrace}\n'
                    f'  Clear();\n'
                    f'  return true;\n'
                    f'{rbrace}\n\n'
                    f'const {dt_name}CfgInfo *{dt_name}Cfg::Get{dt_name}CfgInfo({key_dt} {key_en})\n'
                    f'{lbrace}\n'
                    f'  auto iter = m_{dt_name}CfgMap.find({key_en});\n'
                    f'  return (iter != m_{dt_name}CfgMap.end()) ? &iter->second : nullptr;\n'
                    f'{rbrace}\n')

    except Exception as err:
        if os.path.exists(imp_file_name):
            os.remove(imp_file_name)
        print(f'write_table_class_cpp_file {xlsname}->{sheetname} error : {err} ')
        return False

    print(f'{imp_file_name}')
    return True



# 生成程序代码
def write_program(xlsname, sheet, head_lst, ary_dict):
    ret = write_table_class_h_file(xlsname, sheet.title, head_lst, ary_dict)
    if ret:
        ret = write_table_class_cpp_file(xlsname, sheet.title, head_lst, ary_dict)
    return ret



# 记录所有的表格
g_dict_tablemgr = {}

def add_table_mgr(xlsname, sheetname):
    sheetname_lst = g_dict_tablemgr.get(xlsname)
    if sheetname_lst is None:
       g_dict_tablemgr[xlsname] = [sheetname]
    else:
       if sheetname not in sheetname_lst:
           sheetname_lst.append(sheetname)

#生成 TableMgr.h  TableMgr.cpp 代码
def write_table_mgr_program():
    table_mgr_h_file = f'{h_file_path}TableMgr.h'
    table_mgr_cpp_file = f'{cpp_file_path}TableMgr.cpp'
    lbrace = '{'
    rbrace = '}'
    try:
        with open(table_mgr_h_file, 'w') as f:
            f.write(f'#pragma once\n\n'
                    f'#include <stdint.h>\n'
                    f'#include "base/core/singleton.h"\n\n'
                    f'class TableMgr : public ManualSingleton<TableMgr>\n'
                    f'{lbrace}\n'
                    f'public:\n    '
                    f'TableMgr();\n    '
                    f'~TableMgr();\n    '
                    f'bool Init();\n    '
                    f'bool UnInit();\n'
                    f'{rbrace};\n\n'
                    f'#define  g_MakeTableDataMgr() (TableMgr::Instance())\n'
                    f'#define  g_GetTableDataMgr()  (TableMgr::GetInstance())\n'
                    f'#define  g_DelTableDataMgr()  (TableMgr::Destroy())\n')
    except Exception as err:
        if os.path.exists(table_mgr_h_file):
            os.remove(table_mgr_h_file)
        print(f'write_table_mgr_program....table_mgr_h_file  err:{err} ')
        return False
    print(f'\n{table_mgr_h_file}')
    #
    try:
        with open(table_mgr_cpp_file, 'w') as f:
            include_content = '#include "base/core/os/string.h"\n' \
                              '#include "Common/TableData/TableMgr.h"\n'
            init_content = '    bool result = false;\n    bool retCode = true;\n\n    '
            uninit_content = ''
            for xls, sheetlst in g_dict_tablemgr.items():
                for name in sheetlst:
                    dt_name = f'{xls.capitalize()}{name.capitalize()}'
                    txt_name = f'{xls.capitalize()}_{name}.txt'
                    include_content = f'{include_content}#include "Common/TableData/{dt_name}Cfg.h"\n'
                    init_content = f'{init_content}retCode=g_Make{dt_name}CfgTable()->Init("./data/config/{txt_name}","{split_char}");\n    ' \
                                   f'MMOLOG_PROCESS_ERROR(retCode);\n    '
                    uninit_content = f'{uninit_content}if(g_Get{dt_name}CfgTable())\n    ' \
                                     f'{lbrace}\n    ' \
                                     f'  g_Get{dt_name}CfgTable()->UnInit();\n    ' \
                                     f'  g_Del{dt_name}CfgTable();\n    ' \
                                     f'{rbrace}\n    '
            #
            f.write(f'{include_content}\n\n'
                    f'TableMgr::TableMgr()\n'
                    f'{lbrace}{rbrace}\n\n'
                    f'TableMgr::~TableMgr()\n'
                    f'{lbrace}{rbrace}\n\n'
                    f'bool TableMgr::Init()\n'
                    f'{lbrace}\n'
                    f'{init_content}\n    '
                    f'result = true;\n'
                    f'Exit0:\n    '
                    f'return result;\n'
                    f'{rbrace}\n\n'
                    f'bool TableMgr::UnInit()\n'
                    f'{lbrace}\n    '
                    f'{uninit_content}\n    '
                    f'return true;\n'
                    f'{rbrace}\n')
    except Exception as err:
        print(f'write_table_mgr_program....table_mgr_cpp_file err:{err} ')
        return False
    print(f'{table_mgr_cpp_file}')
    return True

#读取 excel 标签页
def read_sheet(xlsname, sheet, sheetname, trans_type):
    if sheet.max_row < 4:
        return True
    (ret, head_lst, ary_dict ) = read_sheet_head(xlsname, sheet)
    if not ret:
        return False
    if len(head_lst) > 0:  # 表头列数大于0的才需要导出
        ret = write_txt_file(xlsname, sheet, head_lst)
        if ret and trans_type == 0:
            ret = write_program(xlsname, sheet, head_lst, ary_dict)
            if ret:
                add_table_mgr(xlsname, sheet.title)
    return ret




#读取 excel 配置文件
def read_xls(xlsname, trans_type):
    xls = load_workbook(xls_file_path + xlsname + '.xlsx', read_only=False)
    try:
        dict = {}
        for sheet in xls.worksheets:
            sheetname = sheet.title
            if sheetname == 'main':
                continue
            elif sheetname == 'list':
                if sheet.max_row > 0 and sheet.max_column > 0:
                    for r in range(sheet.max_row):
                        val = cell_str_value(sheet, r+1, 1)
                        dict[str(val)] = True
            elif sheetname in dict:
                ret = read_sheet(xlsname, sheet, sheetname, trans_type)
                if not ret:
                    print(f'read_sheet {xlsname}->{sheet.title}  failed...')
                    xls.close()
                    return False
    except Exception as err:
        print(f"open {xlsname}  error: {err} ")
        xls.close()
        return False
    #
    xls.close()
    return True


#转表
def trans_xlsx():
    # 转表类型, 0-转表并生成代码, 1-只转表
    trans_type = 0
    trans_file = []
    result = True
    arglen = len(sys.argv)
    for idx in range(arglen):
        if idx < 1:
            continue
        if idx == 1:
            try:
                trans_type = int(sys.argv[1])
            except Exception as err:
                result = False
                print(f'parse trans_type failed... {err}')
                break
        else:
            file = xls_file_path + sys.argv[idx] + '.xlsx'
            if not os.path.isfile(file):
                result = False
                print(f'{file} not exists...')
                break
            #
            trans_file.append(sys.argv[idx])

    #
    if not result:
        if is_win_system:
            print('按任意键结束')
            msvcrt.getch()
        return
    #
    if len(trans_file) > 0:
        ret = True
        for file in trans_file:
            ret = read_xls(file, trans_type)
            if not ret:
                break
        #
        if ret:
            print('\n\ncongratulations, all excel file process over...\n\n')
    else:
        try:
            with open(list_file, 'r') as list:
                line = next(list)
                line.strip()
                if line[-1:] == '\n':
                    line = line[:-1]
                ret = True
                while line:
                    file = xls_file_path + line + '.xlsx'
                    if not os.path.isfile(file):
                        print(f'{file} not exist')
                        ret = False
                        break
                    ret = read_xls(line, trans_type)
                    if not ret:
                        break
                    try:
                        line = next(list)
                    except:
                        break
                    line.strip()
                    if line[-1:] == '\n':
                        line = line[:-1]
                #
                if ret and trans_type == 0:
                    write_table_mgr_program()
                if ret:
                    print('\n\ncongratulations, all excel file process over...\n\n')
        except:
            print("error:", sys.exc_info()[0])

    if is_win_system:
        print('按任意键结束')
        msvcrt.getch()




trans_xlsx()